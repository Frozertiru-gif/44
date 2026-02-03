from __future__ import annotations

from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
import logging

from aiogram import F, Router
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, CallbackQuery, Message
from openpyxl import Workbook
from sqlalchemy import select

from app.bot.handlers.permissions import FINANCE_EXPORT_ROLES, FINANCE_SUMMARY_ROLES, MANUAL_TX_ROLES, MASTER_ROLES
from app.bot.keyboards.confirmations import confirm_action_keyboard
from app.bot.keyboards.finance import period_keyboard, share_list_keyboard
from app.bot.states.finance import FinanceStates
from app.core.config import get_settings
from app.db.enums import ProjectTransactionType, UserRole, ticket_category_label
from app.domain.enums_mapping import ad_source_label
from app.db.models import User
from app.db.session import async_session_factory
from app.services.audit_service import AuditService
from app.services.finance_service import FinanceService
from app.services.project_settings_service import ProjectSettingsService
from app.services.project_share_service import ProjectShareService
from app.services.project_transaction_service import ProjectTransactionService
from app.services.user_service import UserService

router = Router()
finance_service = FinanceService()
project_transaction_service = ProjectTransactionService()
project_share_service = ProjectShareService()
user_service = UserService()
settings = get_settings()
audit_service = AuditService()
project_settings_service = ProjectSettingsService()
log = logging.getLogger(__name__)


def _parse_amount(value: str) -> Decimal | None:
    cleaned = value.replace(" ", "").replace(",", ".")
    try:
        amount = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    if amount < 0:
        return None
    return amount


def _parse_date(value: str) -> date | None:
    try:
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _period_from_key(key: str) -> tuple[date | None, date | None, str]:
    today = date.today()
    if key == "this_month":
        start = today.replace(day=1)
        return start, today, "Этот месяц"
    if key == "last_month":
        first_this_month = today.replace(day=1)
        last_month_end = first_this_month - timedelta(days=1)
        start = last_month_end.replace(day=1)
        return start, last_month_end, "Прошлый месяц"
    if key == "last_7":
        start = today - timedelta(days=6)
        return start, today, "7 дней"
    if key == "all_time":
        return None, None, "Все время"
    return None, None, "Произвольно"


@router.message(F.text == "💰 Мои деньги")
async def master_money_start(message: Message, state: FSMContext) -> None:
    async with async_session_factory() as session:
        user = await user_service.ensure_user(
            session, message.from_user.id, message.from_user.full_name if message.from_user else None
        )
        await session.commit()
        if not user.is_active or user.role not in MASTER_ROLES:
            await audit_service.log_audit_event(
                session,
                actor_id=user.id,
                action="PERMISSION_DENIED",
                entity_type="finance",
                entity_id=None,
                payload={"reason": "MASTER_MONEY"},
            )
            await session.commit()
            await message.answer("У вас нет доступа к финансам мастера.")
            return
    await state.clear()
    await message.answer("Выберите период:", reply_markup=period_keyboard("finance_master"))


@router.message(F.text == "💵 Моя зарплата")
async def salary_start(message: Message, state: FSMContext) -> None:
    async with async_session_factory() as session:
        user = await user_service.ensure_user(
            session, message.from_user.id, message.from_user.full_name if message.from_user else None
        )
        await session.commit()
        if not user.is_active:
            await message.answer("У вас нет доступа.")
            return
        if user.role not in {UserRole.ADMIN, UserRole.JUNIOR_MASTER, UserRole.SUPER_ADMIN, UserRole.SYS_ADMIN}:
            await audit_service.log_audit_event(
                session,
                actor_id=user.id,
                action="PERMISSION_DENIED",
                entity_type="finance",
                entity_id=None,
                payload={"reason": "SALARY_VIEW"},
            )
            await session.commit()
            await message.answer("У вас нет доступа.")
            return
    await state.clear()
    await message.answer("Выберите период:", reply_markup=period_keyboard("finance_salary"))


@router.message(F.text == "📊 Сводка проекта")
async def project_summary_start(message: Message, state: FSMContext) -> None:
    async with async_session_factory() as session:
        user = await user_service.ensure_user(
            session, message.from_user.id, message.from_user.full_name if message.from_user else None
        )
        await session.commit()
        if not user.is_active or user.role not in FINANCE_SUMMARY_ROLES:
            await audit_service.log_audit_event(
                session,
                actor_id=user.id,
                action="PERMISSION_DENIED",
                entity_type="finance",
                entity_id=None,
                payload={"reason": "FINANCE_SUMMARY"},
            )
            await session.commit()
            await message.answer("У вас нет доступа к сводке.")
            return
    await state.clear()
    await message.answer("Выберите период:", reply_markup=period_keyboard("finance_summary"))


@router.message(F.text == "⬇️ Экспорт Excel")
async def export_start(message: Message, state: FSMContext) -> None:
    async with async_session_factory() as session:
        user = await user_service.ensure_user(
            session, message.from_user.id, message.from_user.full_name if message.from_user else None
        )
        await session.commit()
        if not user.is_active or user.role not in FINANCE_EXPORT_ROLES:
            await audit_service.log_audit_event(
                session,
                actor_id=user.id,
                action="PERMISSION_DENIED",
                entity_type="finance",
                entity_id=None,
                payload={"reason": "FINANCE_EXPORT"},
            )
            await session.commit()
            await message.answer("У вас нет доступа к экспорту.")
            return
    await state.clear()
    await message.answer("Выберите период:", reply_markup=period_keyboard("finance_export"))


@router.callback_query(F.data.startswith("finance_"))
async def finance_period_select(callback: CallbackQuery, state: FSMContext) -> None:
    prefix, period_key = callback.data.split(":", 1)
    flow = prefix.replace("finance_", "")

    if period_key == "custom":
        await state.clear()
        await state.update_data(flow=flow)
        await state.set_state(FinanceStates.period_from)
        await callback.message.answer("Введите дату начала (YYYY-MM-DD):")
        await callback.answer()
        return

    start_date, end_date, label = _period_from_key(period_key)
    await _handle_flow(
        callback.message,
        callback.from_user.id,
        callback.from_user.full_name if callback.from_user else None,
        flow,
        start_date,
        end_date,
        label,
    )
    await callback.answer()


@router.message(FinanceStates.period_from)
async def finance_period_from(message: Message, state: FSMContext) -> None:
    date_value = _parse_date(message.text or "")
    if not date_value:
        await message.answer("Введите дату в формате YYYY-MM-DD.")
        return
    await state.update_data(period_from=date_value)
    await state.set_state(FinanceStates.period_to)
    await message.answer("Введите дату окончания (YYYY-MM-DD):")


@router.message(FinanceStates.period_to)
async def finance_period_to(message: Message, state: FSMContext) -> None:
    date_value = _parse_date(message.text or "")
    if not date_value:
        await message.answer("Введите дату в формате YYYY-MM-DD.")
        return

    data = await state.get_data()
    start_date = data.get("period_from")
    flow = data.get("flow")
    if not isinstance(start_date, date) or not isinstance(flow, str):
        await message.answer("Сессия устарела.")
        await state.clear()
        return
    if date_value < start_date:
        await message.answer("Дата окончания должна быть позже даты начала.")
        return
    await state.clear()
    await _handle_flow(
        message,
        message.from_user.id,
        message.from_user.full_name if message.from_user else None,
        flow,
        start_date,
        date_value,
        "Произвольно",
    )


@router.message(F.text == "➕ Добавить доход")
async def add_income_start(message: Message, state: FSMContext) -> None:
    await _start_transaction_flow(message, state, ProjectTransactionType.INCOME)


@router.message(F.text == "➖ Добавить расход")
async def add_expense_start(message: Message, state: FSMContext) -> None:
    await _start_transaction_flow(message, state, ProjectTransactionType.EXPENSE)


async def _start_transaction_flow(message: Message, state: FSMContext, transaction_type: ProjectTransactionType) -> None:
    async with async_session_factory() as session:
        user = await user_service.ensure_user(
            session, message.from_user.id, message.from_user.full_name if message.from_user else None
        )
        await session.commit()
        if not user.is_active or user.role not in MANUAL_TX_ROLES:
            await audit_service.log_audit_event(
                session,
                actor_id=user.id,
                action="PERMISSION_DENIED",
                entity_type="finance",
                entity_id=None,
                payload={"reason": f"TX_{transaction_type.value}"},
            )
            await session.commit()
            await message.answer("У вас нет доступа к операциям.")
            return
    await state.clear()
    await state.update_data(transaction_type=transaction_type.value)
    await state.set_state(FinanceStates.transaction_amount)
    await message.answer("Введите сумму:")


@router.message(FinanceStates.transaction_amount)
async def transaction_amount(message: Message, state: FSMContext) -> None:
    amount = _parse_amount(message.text or "")
    if amount is None:
        await message.answer("Введите корректную сумму (>= 0).")
        return
    await state.update_data(amount=amount)
    await state.set_state(FinanceStates.transaction_category)
    await message.answer("Введите категорию:")


@router.message(FinanceStates.transaction_category)
async def transaction_category(message: Message, state: FSMContext) -> None:
    category = (message.text or "").strip()
    if not category:
        await message.answer("Категория не может быть пустой.")
        return
    await state.update_data(category=category)
    await state.set_state(FinanceStates.transaction_comment)
    await message.answer("Комментарий (опционально, напишите '-' если без комментария):")


@router.message(FinanceStates.transaction_comment)
async def transaction_comment(message: Message, state: FSMContext) -> None:
    comment = (message.text or "").strip()
    if comment == "-" or not comment:
        comment = None
    await state.update_data(comment=comment)
    await state.set_state(FinanceStates.transaction_date)
    await message.answer("Дата операции (YYYY-MM-DD) или 'сейчас':")


@router.message(FinanceStates.transaction_date)
async def transaction_date(message: Message, state: FSMContext) -> None:
    raw = (message.text or "").strip().lower()
    if raw in {"сейчас", "now", "today", ""}:
        occurred_at = datetime.utcnow()
    else:
        parsed_date = _parse_date(raw)
        if not parsed_date:
            await message.answer("Введите дату в формате YYYY-MM-DD или 'сейчас'.")
            return
        occurred_at = datetime.combine(parsed_date, datetime.min.time())

    data = await state.get_data()
    transaction_type_raw = data.get("transaction_type")
    amount = data.get("amount")
    category = data.get("category")
    comment = data.get("comment")
    if not isinstance(transaction_type_raw, str) or not isinstance(amount, Decimal) or not isinstance(category, str):
        await message.answer("Сессия устарела.")
        await state.clear()
        return

    transaction_type = ProjectTransactionType(transaction_type_raw)

    async with async_session_factory() as session:
        user = await user_service.ensure_user(
            session, message.from_user.id, message.from_user.full_name if message.from_user else None
        )
        if not user.is_active or user.role not in MANUAL_TX_ROLES:
            await message.answer(f"Нет доступа. Ваша роль: {user.role.value}")
            await session.commit()
            await state.clear()
            return

        threshold = await project_settings_service.get_threshold(
            session,
            "large_expense",
            default=10000,
        )
        if transaction_type == ProjectTransactionType.EXPENSE and amount >= Decimal(threshold):
            await state.update_data(occurred_at=occurred_at)
            await state.set_state(FinanceStates.transaction_confirm)
            await message.answer(
                "Вы уверены? Это действие нельзя отменить.",
                reply_markup=confirm_action_keyboard("tx_confirm", "tx_cancel"),
            )
            await session.commit()
            return

        await project_transaction_service.add_transaction(
            session,
            transaction_type=transaction_type,
            amount=amount,
            category=category,
            comment=comment,
            occurred_at=occurred_at,
            created_by=user.id,
        )
        await session.commit()

    await state.clear()
    await message.answer("Операция добавлена.")


@router.callback_query(F.data == "tx_cancel")
async def transaction_cancel(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await callback.answer("Операция отменена")


@router.callback_query(F.data == "tx_confirm")
async def transaction_confirm(callback: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    transaction_type_raw = data.get("transaction_type")
    amount = data.get("amount")
    category = data.get("category")
    comment = data.get("comment")
    occurred_at = data.get("occurred_at")
    if (
        not isinstance(transaction_type_raw, str)
        or not isinstance(amount, Decimal)
        or not isinstance(category, str)
        or not isinstance(occurred_at, datetime)
    ):
        await callback.answer("Сессия устарела", show_alert=True)
        await state.clear()
        return

    transaction_type = ProjectTransactionType(transaction_type_raw)

    async with async_session_factory() as session:
        user = await user_service.ensure_user(
            session, callback.from_user.id, callback.from_user.full_name if callback.from_user else None
        )
        if not user.is_active or user.role not in MANUAL_TX_ROLES:
            await audit_service.log_audit_event(
                session,
                actor_id=user.id,
                action="PERMISSION_DENIED",
                entity_type="finance",
                entity_id=None,
                payload={"reason": f"TX_{transaction_type.value}"},
            )
            await callback.answer(f"Нет доступа. Ваша роль: {user.role.value}", show_alert=True)
            await session.commit()
            await state.clear()
            return

        await project_transaction_service.add_transaction(
            session,
            transaction_type=transaction_type,
            amount=amount,
            category=category,
            comment=comment,
            occurred_at=occurred_at,
            created_by=user.id,
        )
        await session.commit()

    await state.clear()
    await callback.message.answer("Операция добавлена.")
    await callback.answer()


@router.message(F.text == "📌 Доли от кассы")
async def shares_list(message: Message, state: FSMContext) -> None:
    async with async_session_factory() as session:
        actor = await user_service.ensure_user(
            session, message.from_user.id, message.from_user.full_name if message.from_user else None
        )
        await session.commit()
        if not actor.is_active or actor.role not in FINANCE_SUMMARY_ROLES:
            await audit_service.log_audit_event(
                session,
                actor_id=actor.id,
                action="PERMISSION_DENIED",
                entity_type="project_share",
                entity_id=None,
                payload={"reason": "PROJECT_SHARE_LIST"},
            )
            await session.commit()
            await message.answer("У вас нет доступа к долям.")
            return

        users = await user_service.list_users(session, limit=200)
        shares = await finance_service.list_active_shares(session)
    share_map = {share.user_id: share.percent for share in shares}
    entries = []
    lines = ["Доли от кассы:"]
    for user in users:
        percent = share_map.get(user.id)
        percent_label = f"{percent:.2f}%" if percent is not None else "-"
        label = user.display_name or f"ID {user.id}"
        lines.append(f"- {label}: {percent_label}")
        entries.append((user.id, f"{label} ({percent_label})"))

    await state.clear()
    await message.answer("\n".join(lines), reply_markup=share_list_keyboard(entries))


@router.callback_query(F.data.startswith("share_pick:"))
async def share_pick(callback: CallbackQuery, state: FSMContext) -> None:
    user_id = int(callback.data.split(":", 1)[1])
    await state.clear()
    await state.update_data(share_user_id=user_id)
    await state.set_state(FinanceStates.share_percent)
    await callback.message.answer("Введите процент доли (0..100):")
    await callback.answer()


@router.message(FinanceStates.share_percent)
async def share_percent_set(message: Message, state: FSMContext) -> None:
    text = (message.text or "").replace(",", ".").strip()
    try:
        percent = Decimal(text)
    except (InvalidOperation, ValueError):
        await message.answer("Введите корректный процент.")
        return

    data = await state.get_data()
    user_id = data.get("share_user_id")
    if not isinstance(user_id, int):
        await message.answer("Сессия устарела.")
        await state.clear()
        return

    await state.update_data(share_percent=percent)
    await state.set_state(FinanceStates.share_confirm)
    await message.answer(
        "Вы уверены? Это действие нельзя отменить.",
        reply_markup=confirm_action_keyboard("share_confirm", "share_cancel"),
    )


@router.callback_query(F.data == "share_cancel")
async def share_cancel(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await callback.answer("Изменение отменено")


@router.callback_query(F.data == "share_confirm")
async def share_confirm(callback: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    user_id = data.get("share_user_id")
    percent = data.get("share_percent")
    if not isinstance(user_id, int) or not isinstance(percent, Decimal):
        await callback.answer("Сессия устарела", show_alert=True)
        await state.clear()
        return

    async with async_session_factory() as session:
        actor = await user_service.ensure_user(
            session, callback.from_user.id, callback.from_user.full_name if callback.from_user else None
        )
        if not actor.is_active or actor.role not in FINANCE_SUMMARY_ROLES:
            await audit_service.log_audit_event(
                session,
                actor_id=actor.id,
                action="PERMISSION_DENIED",
                entity_type="project_share",
                entity_id=None,
                payload={"reason": "PROJECT_SHARE_SET"},
            )
            await callback.answer(f"Нет доступа. Ваша роль: {actor.role.value}", show_alert=True)
            await session.commit()
            await state.clear()
            return

        try:
            await project_share_service.set_share(
                session,
                user_id=user_id,
                percent=percent,
                actor_id=actor.id,
            )
        except ValueError as exc:
            await callback.answer(str(exc), show_alert=True)
            return
        await session.commit()

    await state.clear()
    await callback.message.answer("Доля обновлена.")
    await callback.answer()


async def _handle_flow(
    message: Message | None,
    tg_user_id: int,
    display_name: str | None,
    flow: str,
    start_date: date | None,
    end_date: date | None,
    label: str,
) -> None:
    if message is None:
        return
    async with async_session_factory() as session:
        actor = await user_service.ensure_user(session, tg_user_id, display_name)
        await session.commit()
        date_range = finance_service.build_range(start_date, end_date)

        if flow == "master":
            if not actor.is_active or actor.role not in MASTER_ROLES:
                await audit_service.log_audit_event(
                    session,
                    actor_id=actor.id,
                    action="PERMISSION_DENIED",
                    entity_type="finance",
                    entity_id=None,
                    payload={"reason": "MASTER_MONEY"},
                )
                await session.commit()
                await message.answer(f"Нет доступа. Ваша роль: {actor.role.value}")
                return
            summary = await finance_service.master_money(session, actor.id, date_range=date_range)
            await message.answer(
                "💰 Мои деньги\n"
                f"Период: {label}\n"
                f"Начислено: {summary['earned']}\n"
                f"Должен перевести: {summary['net_profit']}\n"
                f"Подтверждено: {summary['confirmed']}\n"
                f"Ожидает: {summary['pending']}"
            )
            return

        if flow == "salary":
            if not actor.is_active:
                await audit_service.log_audit_event(
                    session,
                    actor_id=actor.id,
                    action="PERMISSION_DENIED",
                    entity_type="finance",
                    entity_id=None,
                    payload={"reason": "SALARY_VIEW"},
                )
                await session.commit()
                await message.answer(f"Нет доступа. Ваша роль: {actor.role.value}")
                return
            if actor.role == UserRole.ADMIN or actor.role in FINANCE_SUMMARY_ROLES:
                amount = await finance_service.admin_salary(session, actor.id, date_range=date_range)
            elif actor.role == UserRole.JUNIOR_MASTER:
                amount = await finance_service.junior_salary(session, actor.id, date_range=date_range)
            else:
                await audit_service.log_audit_event(
                    session,
                    actor_id=actor.id,
                    action="PERMISSION_DENIED",
                    entity_type="finance",
                    entity_id=None,
                    payload={"reason": "SALARY_VIEW"},
                )
                await session.commit()
                await message.answer(f"Нет доступа. Ваша роль: {actor.role.value}")
                return
            await message.answer(f"💵 Моя зарплата\nПериод: {label}\nСумма: {amount}")
            return

        if flow == "summary":
            if not actor.is_active or actor.role not in FINANCE_SUMMARY_ROLES:
                await audit_service.log_audit_event(
                    session,
                    actor_id=actor.id,
                    action="PERMISSION_DENIED",
                    entity_type="finance",
                    entity_id=None,
                    payload={"reason": "FINANCE_SUMMARY"},
                )
                await session.commit()
                await message.answer(f"Нет доступа. Ваша роль: {actor.role.value}")
                return
            summary = await finance_service.project_summary(session, date_range=date_range)
            await message.answer(
                "📊 Сводка проекта\n"
                f"Период: {label}\n"
                f"Прибыль по заказам (должно быть): {summary['tickets_net_profit_should']}\n"
                f"Прибыль по заказам (получено): {summary['tickets_net_profit_received']}\n"
                f"Ручные доходы: {summary['manual_income_sum']}\n"
                f"Ручные расходы: {summary['manual_expense_sum']}\n"
                f"Общая касса (должно быть): {summary['project_net_cash_should']}\n"
                f"Общая касса (получено): {summary['project_net_cash_received']}\n"
                f"Начислено мастерам: {summary['earned_executor']}\n"
                f"Начислено админам: {summary['earned_admin']}\n"
                f"Начислено младшим мастерам: {summary['earned_junior']}\n"
                f"Остаток проекта: {summary['project_take_sum']}\n"
                f"Закрыто заказов: {summary['closed_count']}\n"
                f"Подтверждено заказов: {summary['confirmed_count']}\n"
                f"Повторов: {summary['repeats_count']}"
            )
            return

        if flow == "export":
            log.info("FINANCE_ACCESS tg=%s actor_id=%s role=%s", tg_user_id, actor.id, actor.role)
            if not actor.is_active or actor.role not in FINANCE_EXPORT_ROLES:
                await audit_service.log_audit_event(
                    session,
                    actor_id=actor.id,
                    action="PERMISSION_DENIED",
                    entity_type="finance",
                    entity_id=None,
                    payload={"reason": "FINANCE_EXPORT"},
                )
                await session.commit()
                await message.answer(f"Нет доступа. Ваша роль: {actor.role.value}")
                return
            tickets = await finance_service.list_tickets_for_export(session, date_range=date_range)
            transactions = await finance_service.list_manual_transactions(session, date_range=date_range)
            summary = await finance_service.project_summary(session, date_range=date_range)
            shares = await finance_service.list_active_shares(session)
            user_map = await _build_user_map(session, tickets, transactions)
            content = _build_excel_report(
                tickets=tickets,
                transactions=transactions,
                summary=summary,
                shares=shares,
                date_range=date_range,
                user_map=user_map,
            )
            filename = f"project_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            target_chat = settings.finance_export_chat_id or actor.id
            await message.bot.send_document(
                chat_id=target_chat,
                document=BufferedInputFile(content.getvalue(), filename=filename),
            )
            await message.answer("Экспорт отправлен.")
            return


async def _build_user_map(session, tickets, transactions) -> dict[int, str]:
    user_ids = set()
    for ticket in tickets:
        if ticket.created_by_admin_id:
            user_ids.add(ticket.created_by_admin_id)
        if ticket.assigned_executor_id:
            user_ids.add(ticket.assigned_executor_id)
        if ticket.junior_master_id:
            user_ids.add(ticket.junior_master_id)
        if ticket.transfer_confirmed_by:
            user_ids.add(ticket.transfer_confirmed_by)
    for tx in transactions:
        user_ids.add(tx.created_by)

    if not user_ids:
        return {}

    result = await session.execute(select(User).where(User.id.in_(user_ids)))
    users = result.scalars().all()
    return {user.id: user.display_name or f"ID {user.id}" for user in users}


def _build_excel_report(*, tickets, transactions, summary, shares, date_range, user_map) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)

    tickets_ws = workbook.create_sheet("Tickets")
    tickets_headers = [
        "ticket_id",
        "status",
        "category",
        "client_phone",
        "scheduled_at",
        "ad_source",
        "created_by_admin",
        "executor",
        "junior_master",
        "revenue",
        "expense",
        "net_profit",
        "transfer_status",
        "transfer_sent_at",
        "transfer_confirmed_at",
        "confirmed_by",
        "is_repeat",
        "repeat_ticket_ids",
    ]
    tickets_ws.append(tickets_headers)

    for ticket in tickets:
        created_by = ticket.created_by.display_name if ticket.created_by else None
        executor = ticket.assigned_executor.display_name if ticket.assigned_executor else None
        junior = ticket.junior_master.display_name if ticket.junior_master else None
        confirmed_by = (
            user_map.get(ticket.transfer_confirmed_by)
            if ticket.transfer_confirmed_by
            else None
        )
        tickets_ws.append(
            [
                ticket.id,
                ticket.status.value if ticket.status else None,
                ticket_category_label(ticket.category) if ticket.category else None,
                ticket.client_phone,
                ticket.scheduled_at,
                ad_source_label(ticket.ad_source) if ticket.ad_source else None,
                created_by or ticket.created_by_admin_id,
                executor or ticket.assigned_executor_id,
                junior or ticket.junior_master_id,
                ticket.revenue,
                ticket.expense,
                ticket.net_profit,
                ticket.transfer_status.value if ticket.transfer_status else None,
                ticket.transfer_sent_at,
                ticket.transfer_confirmed_at,
                confirmed_by or ticket.transfer_confirmed_by,
                ticket.is_repeat,
                ",".join(map(str, ticket.repeat_ticket_ids or [])),
            ]
        )

    report_ws = workbook.create_sheet("OrderReport")
    report_headers = [
        "Номер заказа",
        "Кто выполнил",
        "Тип рекламы",
        "Скок отдал клиент",
        "Расходы",
        "Чистый профит",
    ]
    report_ws.append(report_headers)
    for ticket in tickets:
        executor = ticket.assigned_executor.display_name if ticket.assigned_executor else None
        report_ws.append(
            [
                ticket.id,
                executor or ticket.assigned_executor_id,
                ad_source_label(ticket.ad_source) if ticket.ad_source else None,
                ticket.revenue,
                ticket.expense,
                ticket.net_profit,
            ]
        )

    earnings_ws = workbook.create_sheet("EarningsByTicket")
    earnings_ws.append(
        ["ticket_id", "role", "user_id", "user_name", "percent_at_close", "earned_amount"]
    )
    for ticket in tickets:
        if ticket.assigned_executor_id and ticket.executor_earned_amount is not None:
            name = ticket.assigned_executor.display_name if ticket.assigned_executor else None
            earnings_ws.append(
                [
                    ticket.id,
                    "EXECUTOR",
                    ticket.assigned_executor_id,
                    name,
                    ticket.executor_percent_at_close,
                    ticket.executor_earned_amount,
                ]
            )
        if ticket.created_by_admin_id and ticket.admin_earned_amount is not None:
            name = ticket.created_by.display_name if ticket.created_by else None
            earnings_ws.append(
                [
                    ticket.id,
                    "ADMIN",
                    ticket.created_by_admin_id,
                    name,
                    ticket.admin_percent_at_close,
                    ticket.admin_earned_amount,
                ]
            )
        if ticket.junior_master_id and ticket.junior_master_earned_amount is not None:
            name = ticket.junior_master.display_name if ticket.junior_master else None
            earnings_ws.append(
                [
                    ticket.id,
                    "JUNIOR_MASTER",
                    ticket.junior_master_id,
                    name,
                    ticket.junior_master_percent_at_close,
                    ticket.junior_master_earned_amount,
                ]
            )

    tx_ws = workbook.create_sheet("ManualTransactions")
    tx_ws.append(["id", "type", "amount", "category", "comment", "occurred_at", "created_by"])
    for tx in transactions:
        creator = tx.creator.display_name if tx.creator else None
        tx_ws.append(
            [
                tx.id,
                tx.type.value,
                tx.amount,
                tx.category,
                tx.comment,
                tx.occurred_at,
                creator or tx.created_by,
            ]
        )

    summary_ws = workbook.create_sheet("ProjectSummary")
    summary_ws.append(
        [
            "period_from",
            "period_to",
            "tickets_net_profit_should",
            "tickets_net_profit_received",
            "manual_income_sum",
            "manual_expense_sum",
            "project_net_cash_should",
            "project_net_cash_received",
            "earned_executor",
            "earned_admin",
            "earned_junior",
            "project_take_sum",
            "closed_count",
            "confirmed_count",
            "repeats_count",
        ]
    )
    summary_ws.append(
        [
            date_range.start.date() if date_range.start else None,
            date_range.end.date() if date_range.end else None,
            summary["tickets_net_profit_should"],
            summary["tickets_net_profit_received"],
            summary["manual_income_sum"],
            summary["manual_expense_sum"],
            summary["project_net_cash_should"],
            summary["project_net_cash_received"],
            summary["earned_executor"],
            summary["earned_admin"],
            summary["earned_junior"],
            summary["project_take_sum"],
            summary["closed_count"],
            summary["confirmed_count"],
            summary["repeats_count"],
        ]
    )

    shares_ws = workbook.create_sheet("ProjectShares")
    shares_ws.append(
        [
            "user_id",
            "user_name",
            "percent",
            "project_net_cash_should",
            "share_amount_should",
            "project_net_cash_received",
            "share_amount_received",
        ]
    )
    for share in shares:
        name = share.user.display_name if share.user else None
        share_amount_should = finance_service.round_money(
            summary["project_net_cash_should"] * share.percent / Decimal("100")
        )
        share_amount_received = finance_service.round_money(
            summary["project_net_cash_received"] * share.percent / Decimal("100")
        )
        shares_ws.append(
            [
                share.user_id,
                name,
                share.percent,
                summary["project_net_cash_should"],
                share_amount_should,
                summary["project_net_cash_received"],
                share_amount_received,
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
